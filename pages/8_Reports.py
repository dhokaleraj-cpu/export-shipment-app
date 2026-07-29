from common import *
import html
from reportlab.lib.styles import ParagraphStyle

REPORTS_VERSION = "SN 27.04"

page_setup()
require_page_view("reports")
show_edit_permission_status("reports")

show_header("Reports", "SN 27.04 - Export Shipment Monitoring System")
access_notice()

# ---------------------------------------------------------------------------
# Report list requested for SN 26.00
# ---------------------------------------------------------------------------

REPORTS = [
    "Shipment List",
    "Shipment List with Part Number",
    "Shipment List with Pallet Numbers",
    "Delivery Invoice List against Original Invoice Number",
    "Payment Report with Original Invoice Number",
    "Payment Due Invoice List",
    "Payment Received Report",
    "Customer Wise Shipment Report",
    "Customer Wise Delivery Report",
    "Palletwise Pending Quantity",
]


def _txt(v):
    return str(v or "").strip()


def _like_clause(expr, value):
    value = _txt(value)
    if not value:
        return "", []
    return f" AND LOWER(COALESCE({expr}, '')) LIKE ? ", [f"%{value.lower()}%"]


def _date_clause(expr, from_date, to_date):
    if not from_date or not to_date:
        return "", []
    return f" AND {expr} IS NOT NULL AND {expr}::date BETWEEN ?::date AND ?::date ", [str(from_date), str(to_date)]


def _in_clause(column_name, values):
    values = [int(v) for v in (values or []) if v not in (None, "")]
    if not values:
        return "", []
    return f" AND {column_name} IN ({','.join(['?'] * len(values))}) ", values


def _access_clause(product_column="b.product_id", warehouse_column="s.warehouse_id"):
    clauses, params = [], []
    try:
        product_ids = current_user_allowed_product_ids()
    except Exception:
        product_ids = []
    try:
        warehouse_ids = current_user_allowed_warehouse_ids()
    except Exception:
        warehouse_ids = []

    if product_column and product_ids:
        c, p = _in_clause(product_column, product_ids)
        clauses.append(c)
        params.extend(p)
    if warehouse_column and warehouse_ids:
        c, p = _in_clause(warehouse_column, warehouse_ids)
        clauses.append(c)
        params.extend(p)
    return "".join(clauses), params


def _base_filters(date_expr=None, include_delivery_invoice=True):
    clauses, params = [], []
    for clause, values in [
        _like_clause("s.invoice_no", original_invoice_filter),
        _like_clause("p.product_code || ' ' || p.product_name", part_filter),
        _like_clause("c.customer_name || ' ' || COALESCE(w.warehouse_name,'')", customer_filter),
    ]:
        clauses.append(clause)
        params.extend(values)

    if include_delivery_invoice:
        clause, values = _like_clause("d.delivery_invoice_no", delivery_invoice_filter)
        clauses.append(clause)
        params.extend(values)

        clause, values = _like_clause("d.asn_number", asn_filter)
        clauses.append(clause)
        params.extend(values)

    if date_expr:
        clause, values = _date_clause(date_expr, from_date, to_date)
        clauses.append(clause)
        params.extend(values)

    return "".join(clauses), params


def _shipment_filters(date_expr="s.shipment_date"):
    clauses, params = [], []
    for clause, values in [
        _like_clause("s.invoice_no", original_invoice_filter),
        _like_clause("p.product_code || ' ' || p.product_name", part_filter),
        _like_clause("c.customer_name || ' ' || COALESCE(w.warehouse_name,'')", customer_filter),
        _date_clause(date_expr, from_date, to_date),
    ]:
        clauses.append(clause)
        params.extend(values)
    return "".join(clauses), params


def _run_query(sql, params=None, product_column="b.product_id", warehouse_column="s.warehouse_id"):
    access_sql, access_params = _access_clause(product_column, warehouse_column)
    sql = sql.replace("/*ACCESS_FILTER*/", access_sql)
    final_params = tuple(params or ()) + tuple(access_params)
    if " limit " not in sql.lower():
        sql = sql.rstrip().rstrip(";") + f"\n LIMIT {int(row_limit)}"
    return fetch_all(sql, final_params)


def _format_df(rows):
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(format_date_columns(rows))
    return _format_rate_price_amount_3decimals(df)


def _format_rate_price_amount_3decimals(df):
    """Format rate, price and amount fields to 3 decimals in report grid/export."""
    if df is None or df.empty:
        return df
    out = df.copy()
    decimal_keywords = [
        "rate", "price", "amount", "sale", "value", "paid", "pending", "balance",
        "invoice_amount", "paid_amount", "pending_amount", "payment_received_amount",
        "unit_price", "average_price"
    ]
    for col in out.columns:
        col_l = str(col).lower()
        if any(k in col_l for k in decimal_keywords):
            try:
                out[col] = pd.to_numeric(out[col], errors="coerce").map(lambda x: "" if pd.isna(x) else f"{float(x):.3f}")
            except Exception:
                pass
    return out


def _numeric_sum(series):
    try:
        return pd.to_numeric(series, errors="coerce").fillna(0).sum()
    except Exception:
        return 0


def _add_total_footer(df):
    """Add TOTAL row with correct report footer logic.

    SN 26.10:
    - qty/quantity columns are summed, including pallet_qty, delivery_qty, pending_qty.
    - amount/rate/price/payment/balance columns are summed where applicable.
    - pallet count is calculated only from pallet identifier columns such as pallet_no.
    """
    if df.empty:
        return df

    total_row = {col: "" for col in df.columns}
    total_row[df.columns[0]] = "TOTAL"

    qty_keywords = ["qty", "quantity"]
    amount_keywords = ["amount", "sale", "value", "paid", "pending_amount", "balance_amount", "invoice_amount", "payment_received_amount"]

    for col in df.columns:
        col_l = str(col).lower()
        if any(k in col_l for k in qty_keywords):
            total_row[col] = _numeric_sum(df[col])
        elif any(k in col_l for k in amount_keywords):
            total_row[col] = _numeric_sum(df[col])

    # Pallet count must not replace pallet_qty total.
    for col in df.columns:
        col_l = str(col).lower()
        if col_l in ("pallet_no", "pallet_number", "pallet", "pallet id", "pallet_id"):
            try:
                total_row[col] = f"Count: {df[col].dropna().astype(str).replace('', pd.NA).dropna().nunique()}"
            except Exception:
                pass

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def _safe_number(v):
    try:
        return f"{float(v):,.3f}"
    except Exception:
        return str(v or "")


def _period_text():
    return f"{format_date_ddmmyyyy(from_date)} to {format_date_ddmmyyyy(to_date)}"


def _report_header_html(title):
    logo_html = ""
    try:
        if LOGO_PATH.exists():
            logo_b64 = base64.b64encode(LOGO_PATH.read_bytes()).decode("utf-8")
            logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:54px;max-width:280px;object-fit:contain;">'
    except Exception:
        logo_html = ""

    return f"""
    <div style="width:100%;box-sizing:border-box;border:1px solid #CBD5E1;border-radius:12px;padding:12px 16px;margin:10px 0 12px 0;background:white;display:grid;grid-template-columns:1.2fr 2.2fr 1.2fr;align-items:center;gap:12px;box-shadow:0 2px 8px rgba(15,23,42,.06);">
        <div style="text-align:left;">{logo_html}</div>
        <div style="text-align:center;font-size:24px;font-weight:900;color:#003B73;line-height:1.15;">{title}</div>
        <div style="text-align:right;font-size:14px;font-weight:900;color:#334155;line-height:1.35;">Report Period<br>{_period_text()}</div>
    </div>
    """

def _report_footer_totals(df):
    """Return footer totals for report KPI cards.

    SN 26.11:
    - No raw HTML footer is returned, preventing HTML code from showing on screen.
    - Palletwise Pending Quantity shows separate Pallet Qty, Delivery Qty, Pending Qty and Pallet Count.
    """
    result = {
        "qty_total": 0,
        "amount_total": 0,
        "pallet_count": "",
        "pallet_qty": None,
        "delivery_qty": None,
        "pending_qty": None,
    }
    if df.empty:
        return result

    qty_columns = []
    amount_columns = []
    pallet_id_columns = []

    for col in df.columns:
        cl = str(col).lower()
        if "qty" in cl or "quantity" in cl:
            qty_columns.append(col)
        if any(k in cl for k in ["amount", "sale", "value", "paid", "pending_amount", "balance_amount", "invoice_amount", "payment_received_amount"]):
            amount_columns.append(col)
        if cl in ("pallet_no", "pallet_number", "pallet", "pallet id", "pallet_id"):
            pallet_id_columns.append(col)

    result["qty_total"] = sum(_numeric_sum(df[col]) for col in qty_columns)
    result["amount_total"] = sum(_numeric_sum(df[col]) for col in amount_columns)

    if pallet_id_columns:
        col = pallet_id_columns[0]
        try:
            result["pallet_count"] = str(df[col].dropna().astype(str).replace("", pd.NA).dropna().nunique())
        except Exception:
            result["pallet_count"] = ""

    for special_col in ["pallet_qty", "delivery_qty", "pending_qty"]:
        actual_col = next((c for c in df.columns if str(c).lower() == special_col), None)
        if actual_col:
            result[special_col] = _numeric_sum(df[actual_col])

    return result


def _render_report_footer_kpis(df):
    """Render report footer totals as Streamlit cards, not HTML text."""
    if df.empty:
        return

    totals = _report_footer_totals(df)

    st.markdown(
        """
        <style>
        .sn-report-footer-title {
            background:#003B73;
            color:white;
            padding:8px 12px;
            border-radius:12px 12px 0 0;
            font-size:16px;
            font-weight:900;
            text-transform:uppercase;
            letter-spacing:.3px;
            margin-top:12px;
        }
        .sn-report-kpi {
            border:1px solid #BFD7F0;
            background:#EAF3FC;
            padding:12px;
            text-align:center;
            min-height:78px;
        }
        .sn-report-kpi-label {
            font-size:13px;
            font-weight:900;
            color:#003B73;
            text-transform:uppercase;
            margin-bottom:5px;
        }
        .sn-report-kpi-value {
            font-size:20px;
            font-weight:900;
            color:#111827;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="sn-report-footer-title">Report Footer Totals</div>', unsafe_allow_html=True)

    if totals.get("pallet_qty") is not None or totals.get("delivery_qty") is not None or totals.get("pending_qty") is not None:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Total Pallet Qty</div><div class="sn-report-kpi-value">{_safe_number(totals.get("pallet_qty") or 0)}</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Total Delivery Qty</div><div class="sn-report-kpi-value">{_safe_number(totals.get("delivery_qty") or 0)}</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Total Pending Qty</div><div class="sn-report-kpi-value">{_safe_number(totals.get("pending_qty") or 0)}</div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Pallet Count</div><div class="sn-report-kpi-value">{totals.get("pallet_count") or ""}</div></div>', unsafe_allow_html=True)
    else:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Total Qty</div><div class="sn-report-kpi-value">{_safe_number(totals.get("qty_total") or 0)}</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Total Amount</div><div class="sn-report-kpi-value">{_safe_number(totals.get("amount_total") or 0)}</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="sn-report-kpi"><div class="sn-report-kpi-label">Pallet Count</div><div class="sn-report-kpi-value">{totals.get("pallet_count") or ""}</div></div>', unsafe_allow_html=True)


def _excel_bytes(df, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        startrow = 5
        df_export = _format_rate_price_amount_3decimals(_add_total_footer(df.copy()))
        df_export.to_excel(writer, index=False, sheet_name="Report", startrow=startrow)
        ws = writer.sheets["Report"]

        # Report header on Excel sheet.
        ws["A1"] = ""
        ws["B1"] = title
        ws["B2"] = f"Report Period: {_period_text()}"
        ws["B1"].font = ws["B1"].font.copy(bold=True, size=16)
        ws["B2"].font = ws["B2"].font.copy(bold=True, size=11)

        try:
            from openpyxl.drawing.image import Image as XLImage
            if LOGO_PATH.exists():
                img = XLImage(str(LOGO_PATH))
                img.height = 55
                img.width = 230
                ws.add_image(img, "A1")
        except Exception:
            pass

        # Style table header and footer.
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_row = startrow + 1
            total_row = header_row + len(df_export)
            blue_fill = PatternFill("solid", fgColor="003B73")
            total_fill = PatternFill("solid", fgColor="EAF3FC")
            white_font = Font(color="FFFFFF", bold=True)
            bold_blue = Font(color="003B73", bold=True)
            thin = Side(style="thin", color="CBD5E1")
            for cell in ws[header_row]:
                cell.fill = blue_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for cell in ws[total_row]:
                cell.fill = total_fill
                cell.font = bold_blue
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        except Exception:
            pass

        # Repeat report header and column header on printed pages.
        try:
            ws.print_title_rows = f"1:{startrow+1}"
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
        except Exception:
            pass

        for column_cells in ws.columns:
            try:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 34)
            except Exception:
                pass
    return output.getvalue()

def _pdf_bytes(df, title):
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=86,
        bottomMargin=34,
    )
    styles = getSampleStyleSheet()

    def _draw_page_header_footer(canvas, doc_obj):
        canvas.saveState()
        page_width, page_height = landscape(A4)

        # Header box matching report/table width.
        x = 18
        y = page_height - 72
        width = page_width - 36
        height = 54
        canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
        canvas.setFillColor(colors.white)
        canvas.roundRect(x, y, width, height, 6, stroke=1, fill=1)

        # Logo at left.
        try:
            if LOGO_PATH.exists():
                canvas.drawImage(str(LOGO_PATH), x + 10, y + 8, width=145, height=38, preserveAspectRatio=True, mask='auto')
        except Exception:
            canvas.setFillColor(colors.HexColor("#003B73"))
            canvas.setFont("Helvetica-Bold", 12)
            canvas.drawString(x + 12, y + 22, "FSI")

        # Title middle.
        canvas.setFillColor(colors.HexColor("#003B73"))
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_width / 2, y + 29, str(title)[:95])

        # Period right.
        canvas.setFillColor(colors.HexColor("#334155"))
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawRightString(page_width - 28, y + 33, "Report Period")
        canvas.drawRightString(page_width - 28, y + 20, _period_text())

        # Footer with page number.
        canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
        canvas.line(18, 24, page_width - 18, 24)
        canvas.setFillColor(colors.HexColor("#003B73"))
        canvas.setFont("Helvetica-Bold", 8)
        canvas.drawString(18, 12, "Export Shipment Monitoring System")
        canvas.drawCentredString(page_width / 2, 12, f"{title}")
        canvas.drawRightString(page_width - 18, 12, f"Page {doc_obj.page}")

        canvas.restoreState()

    story = []

    pdf_df = _format_rate_price_amount_3decimals(_add_total_footer(df.copy()))
    wrap_style = ParagraphStyle("sn_report_wrap", parent=styles["Normal"], fontName="Helvetica", fontSize=5.6, leading=6.2, wordWrap="CJK")
    head_style = ParagraphStyle("sn_report_head_wrap", parent=wrap_style, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

    def _wrap_pdf_cell(value, style):
        try:
            safe = html.escape(str(value or "")).replace("\n", "<br/>")
            return Paragraph(safe, style)
        except Exception:
            return str(value or "")

    data = [[_wrap_pdf_cell(c, head_style) for c in list(pdf_df.columns)]]
    for row in pdf_df.astype(str).values.tolist():
        data.append([_wrap_pdf_cell(cell, wrap_style) for cell in row])

    # Force table to use full available page width, matching the report header.
    available_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
    col_count = max(len(pdf_df.columns), 1)
    col_widths = [available_width / col_count] * col_count
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#003B73")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 6),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 2),
        ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#EAF3FC")),
        ("TEXTCOLOR", (0,-1), (-1,-1), colors.HexColor("#003B73")),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
    ]))
    story.append(table)
    doc.build(story, onFirstPage=_draw_page_header_footer, onLaterPages=_draw_page_header_footer)
    return output.getvalue()

def _display_report(rows, title):
    df = _format_df(rows)
    st.markdown(_report_header_html(title), unsafe_allow_html=True)

    if df.empty:
        st.info("No records found for selected filters.")
        return

    _render_report_footer_kpis(df)
    st.markdown("<style>/* SN2603 full width report table */ div[data-testid='stDataFrame']{width:100% !important;} div[data-testid='stDataFrame'] > div{width:100% !important;}</style>", unsafe_allow_html=True)
    st.dataframe(_format_rate_price_amount_3decimals(_add_total_footer(df)), width="stretch", hide_index=True)

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Export Excel",
            data=_excel_bytes(df, title),
            file_name=f"{title.replace(' ', '_').lower()}_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    with c2:
        st.download_button(
            "Export PDF",
            data=_pdf_bytes(df, title),
            file_name=f"{title.replace(' ', '_').lower()}_{date.today().isoformat()}.pdf",
            mime="application/pdf",
            width="stretch",
        )


def get_report_rows(report_name):
    if report_name == "Shipment List":
        fsql, params = _shipment_filters("s.shipment_date")
        return _run_query(f"""
            SELECT
                s.shipment_no,
                s.shipment_date,
                s.po_number AS bl_number,
                w.warehouse_name,
                c.customer_name,
                s.invoice_no AS original_invoice_no,
                SUM(b.amount) AS amount
            FROM shipments s
            JOIN shipment_boxes b ON b.shipment_id = s.id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = s.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY s.shipment_no, s.shipment_date, s.po_number, w.warehouse_name, c.customer_name, s.invoice_no
            ORDER BY s.shipment_date DESC, s.shipment_no
        """, params)

    if report_name == "Shipment List with Part Number":
        fsql, params = _shipment_filters("s.shipment_date")
        return _run_query(f"""
            SELECT
                s.shipment_no,
                s.shipment_date,
                s.invoice_no AS original_invoice_no,
                p.product_code,
                p.product_name,
                SUM(b.original_qty) AS qty,
                SUM(b.amount) AS amount
            FROM shipments s
            JOIN shipment_boxes b ON b.shipment_id = s.id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = s.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY s.shipment_no, s.shipment_date, s.invoice_no, p.product_code, p.product_name
            ORDER BY s.shipment_date DESC, s.shipment_no, p.product_code
        """, params)

    if report_name == "Shipment List with Pallet Numbers":
        fsql, params = _shipment_filters("s.shipment_date")
        return _run_query(f"""
            SELECT
                s.shipment_no,
                s.shipment_date,
                s.invoice_no AS original_invoice_no,
                b.pallet_no,
                b.box_no,
                p.product_code,
                p.product_name,
                b.original_qty AS qty
            FROM shipments s
            JOIN shipment_boxes b ON b.shipment_id = s.id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = s.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            ORDER BY s.shipment_date DESC, s.shipment_no, b.pallet_no, b.box_no
        """, params)

    if report_name == "__REMOVED_Delivery Invoice List with Original Invoice Number":
        fsql, params = _base_filters("d.delivery_date")
        return _run_query(f"""
            SELECT
                d.delivery_invoice_no,
                d.delivery_date,
                s.invoice_no AS original_invoice_no,
                p.product_code,
                p.product_name,
                b.pallet_no,
                b.box_no,
                SUM(d.delivered_qty) AS qty,
                AVG(d.unit_price) AS price,
                SUM(d.sale_amount) AS amount
            FROM customer_deliveries d
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY d.delivery_invoice_no, d.delivery_date, s.invoice_no, p.product_code, p.product_name, b.pallet_no, b.box_no
            ORDER BY d.delivery_date DESC, d.delivery_invoice_no
        """, params)

    if report_name == "Delivery Invoice List against Original Invoice Number":
        fsql, params = _base_filters("d.delivery_date")
        return _run_query(f"""
            SELECT
                s.invoice_no AS original_invoice_no,
                d.delivery_invoice_no,
                d.asn_number,
                d.delivery_date,
                c.customer_name,
                p.product_code,
                p.product_name,
                SUM(d.delivered_qty) AS qty,
                SUM(d.sale_amount) AS amount
            FROM customer_deliveries d
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY s.invoice_no, d.delivery_invoice_no, d.asn_number, d.delivery_date, c.customer_name, p.product_code, p.product_name
            ORDER BY s.invoice_no, d.delivery_date DESC, d.delivery_invoice_no
        """, params)

    if report_name == "Payment Report with Original Invoice Number":
        fsql, params = _base_filters("COALESCE(pay.payment_received_date, d.payment_due_date)")
        return _run_query(f"""
            SELECT
                s.invoice_no AS original_invoice_no,
                d.delivery_invoice_no,
                d.asn_number,
                c.customer_name,
                p.product_code,
                p.product_name,
                d.payment_due_date,
                SUM(d.sale_amount) AS invoice_amount,
                COALESCE(SUM(pay.payment_amount),0) AS paid_amount,
                SUM(d.sale_amount) - COALESCE(SUM(pay.payment_amount),0) AS pending_amount
            FROM customer_deliveries d
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            LEFT JOIN payments pay ON pay.delivery_id = d.id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY s.invoice_no, d.delivery_invoice_no, d.asn_number, c.customer_name, p.product_code, p.product_name, d.payment_due_date
            ORDER BY d.payment_due_date, s.invoice_no, d.delivery_invoice_no
        """, params)

    if report_name == "Payment Due Invoice List":
        fsql, params = _base_filters("d.payment_due_date")
        return _run_query(f"""
            SELECT
                s.invoice_no AS original_invoice_no,
                d.delivery_invoice_no,
                d.asn_number,
                c.customer_name,
                d.delivery_date,
                d.payment_due_date,
                SUM(d.sale_amount) AS invoice_amount,
                COALESCE(SUM(pay.payment_amount),0) AS paid_amount,
                SUM(d.sale_amount) - COALESCE(SUM(pay.payment_amount),0) AS pending_amount,
                CASE
                    WHEN SUM(d.sale_amount) - COALESCE(SUM(pay.payment_amount),0) <= 0 THEN 'Paid'
                    WHEN MAX(d.payment_due_date)::date < CURRENT_DATE THEN 'Overdue'
                    ELSE 'Pending'
                END AS payment_status
            FROM customer_deliveries d
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            LEFT JOIN payments pay ON pay.delivery_id = d.id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY s.invoice_no, d.delivery_invoice_no, d.asn_number, c.customer_name, d.delivery_date, d.payment_due_date
            HAVING SUM(d.sale_amount) - COALESCE(SUM(pay.payment_amount),0) > 0
            ORDER BY d.payment_due_date, d.delivery_invoice_no
        """, params)

    if report_name == "Payment Received Report":
        fsql, params = _base_filters("pay.payment_received_date")
        return _run_query(f"""
            SELECT
                pay.payment_received_date,
                pay.payment_reference,
                s.invoice_no AS original_invoice_no,
                d.delivery_invoice_no,
                c.customer_name,
                p.product_code,
                p.product_name,
                SUM(pay.payment_amount) AS payment_received_amount
            FROM payments pay
            JOIN customer_deliveries d ON d.id = pay.delivery_id
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY pay.payment_received_date, pay.payment_reference, s.invoice_no, d.delivery_invoice_no, d.asn_number, c.customer_name, p.product_code, p.product_name
            ORDER BY pay.payment_received_date DESC, pay.payment_reference
        """, params)

    if report_name == "Customer Wise Shipment Report":
        fsql, params = _shipment_filters("s.shipment_date")
        return _run_query(f"""
            SELECT
                c.customer_name,
                w.warehouse_name,
                s.shipment_no,
                s.shipment_date,
                s.invoice_no AS original_invoice_no,
                p.product_code,
                p.product_name,
                SUM(b.original_qty) AS qty,
                SUM(b.amount) AS amount
            FROM shipments s
            JOIN shipment_boxes b ON b.shipment_id = s.id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = s.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY c.customer_name, w.warehouse_name, s.shipment_no, s.shipment_date, s.invoice_no, p.product_code, p.product_name
            ORDER BY c.customer_name, s.shipment_date DESC, s.shipment_no
        """, params)

    if report_name == "Customer Wise Delivery Report":
        fsql, params = _base_filters("d.delivery_date")
        return _run_query(f"""
            SELECT
                c.customer_name,
                d.delivery_invoice_no,
                d.asn_number,
                d.delivery_date,
                s.invoice_no AS original_invoice_no,
                p.product_code,
                p.product_name,
                SUM(d.delivered_qty) AS qty,
                SUM(d.sale_amount) AS amount
            FROM customer_deliveries d
            JOIN shipment_boxes b ON b.id = d.box_id
            JOIN shipments s ON s.id = d.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = d.customer_id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            GROUP BY c.customer_name, d.delivery_invoice_no, d.asn_number, d.delivery_date, s.invoice_no, p.product_code, p.product_name
            ORDER BY c.customer_name, d.delivery_date DESC, d.delivery_invoice_no
        """, params)

    if report_name == "Palletwise Pending Quantity":
        clauses, params = [], []
        for clause, values in [
            _like_clause("s.invoice_no", original_invoice_filter),
            _like_clause("p.product_code || ' ' || p.product_name", part_filter),
            _like_clause("c.customer_name || ' ' || COALESCE(w.warehouse_name,'')", customer_filter),
            _date_clause("s.shipment_date", from_date, to_date),
        ]:
            clauses.append(clause)
            params.extend(values)
        fsql = "".join(clauses)
        return _run_query(f"""
            SELECT
                s.invoice_no AS original_invoice_no,
                s.shipment_no,
                s.shipment_date,
                c.customer_name,
                w.warehouse_name,
                p.product_code,
                p.product_name,
                b.pallet_no,
                b.box_no,
                b.original_qty AS pallet_qty,
                COALESCE(del.delivery_qty,0) AS delivery_qty,
                b.original_qty - COALESCE(del.delivery_qty,0) AS pending_qty
            FROM shipment_boxes b
            JOIN shipments s ON s.id = b.shipment_id
            JOIN products p ON p.id = b.product_id
            LEFT JOIN warehouses w ON w.id = s.warehouse_id
            LEFT JOIN customers c ON c.id = s.customer_id
            LEFT JOIN (
                SELECT box_id, SUM(delivered_qty) AS delivery_qty
                FROM customer_deliveries
                GROUP BY box_id
            ) del ON del.box_id = b.id
            WHERE 1=1
            {fsql}
            /*ACCESS_FILTER*/
            ORDER BY s.shipment_date DESC, s.invoice_no, b.pallet_no, b.box_no, p.product_code
        """, params)

    return []


st.markdown(
    """
    <div class="sap-grid-card">
        <div class="sap-grid-card-title">REPORT FILTERS - SN 26.00</div>
    """,
    unsafe_allow_html=True,
)

r1, r2, r3 = st.columns([2, 1, 1])
with r1:
    selected_report = searchable_selectbox("Select Report", REPORTS, key="sn26_report_select")
with r2:
    from_date = st.date_input("From Date", value=date(2025, 1, 1), key="sn26_from_date")
with r3:
    to_date = st.date_input("To Date", value=date.today(), key="sn26_to_date")

f1, f2, f3, f4, f5 = st.columns([1.2, 1.2, 1.2, 1.2, .8])
with f1:
    original_invoice_filter = st.text_input("Original Invoice Number", key="sn26_original_invoice_filter")
with f2:
    part_filter = st.text_input("Part Number", key="sn26_part_filter")
with f3:
    customer_filter = st.text_input("Customer", key="sn26_customer_filter")
with f4:
    delivery_invoice_filter = st.text_input("Delivery Invoice Number", key="sn26_delivery_invoice_filter")
with f5:
    asn_filter = st.text_input("ASN Number", key="sn26_asn_filter")

row_limit = st.selectbox("Max Rows", [100, 250, 500, 1000, 2000, 5000], index=2, key="sn26_row_limit")

generate = st.button("GENERATE REPORT", type="primary", width="stretch", key="sn26_generate_report")
st.markdown("</div>", unsafe_allow_html=True)

if from_date > to_date:
    st.warning("From Date is after To Date. Dates were swapped.")
    from_date, to_date = to_date, from_date

if not generate:
    st.info("Select report and filters, then click GENERATE REPORT.")
    render_slogan_footer()
    st.stop()

try:
    rows = get_report_rows(selected_report)
    _display_report(rows, selected_report)
except Exception as exc:
    st.error(f"Report could not load: {exc}")
    st.exception(exc)

render_slogan_footer()
